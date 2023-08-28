import openpyxl, getpass, imaplib, sys, time
from email.message import EmailMessage

if len(sys.argv) != 4:
    print("usage: {} user excel body".format(sys.argv[0]))
    sys.exit(1)

server = 'imap.gmail.com'
user = sys.argv[1]
excel = sys.argv[2]
# subject = sys.argv[3]
# sender = sys.argv[4]
mailtxt = sys.argv[3]

print("server: {}".format(server))
print("user: {}".format(user))
print("excel file: {}".format(excel))

skel = ''
sender = ''
subject = ''
with open(mailtxt, 'r') as f:
    for line in f:
        d = line.rstrip()
        if (d.split(' ')[0] == 'From:'):
            sender = d.replace('From: ', '', 1)
        elif (d.split(' ')[0] == 'Subject:'):
            subject = d.replace('Subject: ', '', 1)
        else:
            skel = skel + line
if sender == '':
    print('Error: From: not fund')
    sys.exit(1)
if subject == '':
    print('Error: Subject: not fund')
    sys.exit(1)

print("---")
print("Subject: {}".format(subject))
print("From: {}".format(sender))
print(skel)

book = openpyxl.load_workbook(excel)
sheet = book.active
col = {}
ccs = []
for c in range(1,sheet.max_column+1):
    v = sheet.cell(column=c, row=1).value
    if v != None:
        if v == 'cc':
            ccs.append(c)
        else:
            col[v] = c
print("columns:", col)
print("cc fields:", ccs)

# login
passwd = getpass.getpass()
m = imaplib.IMAP4_SSL(server, '993')
m.login(user, passwd)

# choose draft folder
type, data = m.list()
for d in data:
    d = d.decode('utf-8').split(' "/" ')
    if d[0] == '(\\Drafts \\HasNoChildren)':
        folder = d[1].strip('"')
if folder != "":
    print("draft folder: {}".format(folder))
else:
    print("draft folder not found")
    sys.exit(1)

type, data = m.select(folder)
for r in range(2,sheet.max_row+1):
    format = {k : sheet.cell(column=col[k], row=r).value for k in col.keys()}
    if 'to' in format:
        print(format)
        msg = EmailMessage()
        msg['Subject'] = subject.format(**format)
        msg['From'] = sender.format(**format)
        if 'name' in format:
            to = "{name} <{to}>".format(**format)
        else:
            to = "{to}".format(**format)
        print("To:", to)
        msg['To'] = to
        cc = ", ".join([ sheet.cell(column=c, row=r).value for c in ccs if sheet.cell(column=c, row=r).value != None])
        if cc != "":
            print("cc:", cc)
            msg['Cc'] = cc
        body = skel.format(**format)
        msg.set_content(body)
        m.append(folder, '', imaplib.Time2Internaldate(time.time()), str(msg).encode('utf-8'))
m.close()
m.logout()
